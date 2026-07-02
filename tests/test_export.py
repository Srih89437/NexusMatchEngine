from unittest.mock import patch, MagicMock
from fastapi.testclient import TestClient
import openpyxl
from io import BytesIO
from src.member4_orchestration.main import app

client = TestClient(app)


@patch("src.member4_orchestration.main.get_llm_ranker")
@patch("src.member4_orchestration.main.get_postgres_client")
@patch("src.member4_orchestration.main.get_qdrant_client")
@patch("src.member4_orchestration.main.get_embedder")
def test_xlsx_export_endpoint(
    mock_get_embedder, mock_get_qdrant, mock_get_pg, mock_get_llm
):
    # Setup mock behaviors
    mock_embedder_inst = MagicMock()
    mock_embedder_inst.generate_embeddings.return_value = [
        {"dense": [0.1] * 1024, "sparse": {"indices": [101], "values": [0.5]}}
    ]
    mock_get_embedder.return_value = mock_embedder_inst

    mock_qdrant_inst = MagicMock()
    mock_qdrant_inst.hybrid_search.return_value = [
        {
            "id": "cand_001",
            "score": 0.892,
            "payload": {"name": "Jane Doe", "skills": ["Python", "FastAPI"]},
        }
    ]
    mock_get_qdrant.return_value = mock_qdrant_inst

    mock_pg_inst = MagicMock()
    # Mocking get_job_description to find the job description successfully
    mock_pg_inst.get_job_description.return_value = {
        "id": "job_101",
        "title": "Senior Engineer",
        "required_skills": ["Python", "FastAPI"],
        "min_experience_years": 5,
        "full_text": "Need a senior backend engineer skilled in Python and FastAPI",
    }
    mock_pg_inst.get_candidate.return_value = {
        "id": "cand_001",
        "name": "Jane Doe",
        "skills": ["Python", "FastAPI"],
        "experience": [
            {"company": "Google", "role": "Engineer", "duration_months": 24}
        ],
        "education": [
            {
                "institution": "Stanford",
                "degree": "M.S.",
                "field_of_study": "CS",
                "start_year": 2018,
                "end_year": 2020,
            }
        ],
    }
    mock_get_pg.return_value = mock_pg_inst

    mock_llm_inst = MagicMock()
    mock_llm_inst.rerank_list.return_value = [
        {
            "id": "cand_001",
            "name": "Jane Doe",
            "ltr_score": 0.82,
            "listwise_rank": 1,
            "llm_rationale": "Perfect fit.",
        }
    ]
    mock_get_llm.return_value = mock_llm_inst

    # Make request to the new endpoint
    response = client.get("/api/v1/results/job_101/download")
    assert response.status_code == 200
    assert (
        response.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "content-disposition" in response.headers
    assert (
        "attachment; filename=nexusmatch_results_job_101.xlsx"
        in response.headers["content-disposition"]
    )

    # Verify that the downloaded file is a valid openpyxl Workbook
    wb = openpyxl.load_workbook(BytesIO(response.content))
    assert "Ranked Candidates" in wb.sheetnames
    ws = wb["Ranked Candidates"]

    # Assert headers
    headers = [cell.value for cell in ws[1]]
    expected_headers = [
        "Rank",
        "Candidate ID",
        "Candidate Name",
        "Initial Retrieval Score",
        "LTR Score",
        "Skills",
        "Experience",
        "Education",
        "SHAP Summary / Important Features",
        "LLM Explanation",
    ]
    assert headers == expected_headers

    # Assert row values
    row_values = [cell.value for cell in ws[2]]
    assert row_values[0] == 1  # Rank
    assert row_values[1] == "cand_001"  # Candidate ID
    assert row_values[2] == "Jane Doe"  # Name
    assert row_values[9] == "Perfect fit."  # LLM Explanation
